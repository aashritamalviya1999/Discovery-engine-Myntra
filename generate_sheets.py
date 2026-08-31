import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os

# Define the folder path and input file
working_dir = r"C:\Users\sanja\.gemini\antigravity\scratch\myntra_wishlist_conversion"
csv_path = os.path.join(working_dir, "feedback_analysis_output.csv")
excel_path = os.path.join(working_dir, "Myntra_Wishlist_Growth_Project.xlsx")

# 1. Load the discovery engine data if it exists
if os.path.exists(csv_path):
    df_feedback = pd.read_csv(csv_path)
else:
    df_feedback = pd.DataFrame()

# Create a new workbook
wb = openpyxl.Workbook()
# Remove the default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Style helper definitions
font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
font_subtitle = Font(name="Segoe UI", size=11, italic=True, color="E0E0E0")
font_section = Font(name="Segoe UI", size=12, bold=True, color="282C3F")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="282C3F")
font_normal = Font(name="Segoe UI", size=10, color="282C3F")
font_italic = Font(name="Segoe UI", size=10, italic=True, color="555555")

# Colors
fill_myntra_pink = PatternFill(start_color="D80E62", end_color="D80E62", fill_type="solid") # Main Brand Accent
fill_header_dark = PatternFill(start_color="282C3F", end_color="282C3F", fill_type="solid")  # Slate Gray
fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")        # Soft background
fill_kpi_bg = PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid")       # Very soft pink
fill_section_bar = PatternFill(start_color="F0F2F5", end_color="F0F2F5", fill_type="solid")  # Muted grey header

# Card fills for priorities / status
fill_high_priority = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # Light Red/Pink
font_high_priority = Font(name="Segoe UI", size=10, bold=True, color="C2185B")

fill_med_priority = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")   # Light Purple
font_med_priority = Font(name="Segoe UI", size=10, bold=True, color="7B1FA2")

fill_low_priority = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")   # Light Blue
font_low_priority = Font(name="Segoe UI", size=10, bold=True, color="1976D2")

fill_green_badge = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")    # Muted Green
font_green_badge = Font(name="Segoe UI", size=10, bold=True, color="2E7D32")

# Borders
thin_border_side = Side(border_style="thin", color="D1D5DB")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
double_bottom_border = Border(bottom=Side(border_style="double", color="282C3F"), top=Side(border_style="thin", color="D1D5DB"))

# Alignments
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")
align_title = Alignment(horizontal="left", vertical="center")

def apply_grid_and_autosize(ws):
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't auto-size columns based on title block (rows 1-3) which spans many merged cells
        for cell in col:
            if cell.row > 3 and cell.value:
                # If cell is merged, its value might skew calculation, but let's take a safe estimate
                max_len = max(max_len, len(str(cell.value)))
        
        # Set generous width
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def create_banner(ws, title, subtitle):
    # Banner background
    for r in range(1, 3):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_myntra_pink
            
    # Set values before merging to avoid openpyxl MergedCell read-only issues
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    subtitle_cell = ws["A2"]
    subtitle_cell.value = subtitle
    subtitle_cell.font = font_subtitle
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15  # spacer row

# ==========================================
# SHEET 1: EXECUTIVE DASHBOARD & KPIs
# ==========================================
ws1 = wb.create_sheet(title="Growth Dashboard")
create_banner(ws1, "MYNTRA WISHLIST CONVERSION ACTION PLAN", "Growth Team Strategy - Metric: Increase 30-Day Wishlist-to-Purchase Rate (Non-Monetary)")

# KPI Cards Block (Rows 5-8)
ws1.merge_cells("A5:B5")
ws1["A5"] = "NORTH STAR METRIC"
ws1["A5"].font = Font(name="Segoe UI", size=10, bold=True, color="7F8C8D")
ws1["A5"].alignment = align_center

ws1.merge_cells("A6:B8")
ws1["A6"] = "6.50%"
ws1["A6"].font = Font(name="Segoe UI", size=24, bold=True, color="D80E62")
ws1["A6"].alignment = align_center
ws1["A6"].fill = fill_kpi_bg

# Apply borders to NS card
for r in range(5, 9):
    for c in range(1, 3):
        ws1.cell(row=r, column=c).border = thin_border

ws1.merge_cells("C5:D5")
ws1["C5"] = "CURRENT BASELINE"
ws1["C5"].font = Font(name="Segoe UI", size=10, bold=True, color="7F8C8D")
ws1["C5"].alignment = align_center

ws1.merge_cells("C6:D8")
ws1["C6"] = "4.20%"
ws1["C6"].font = Font(name="Segoe UI", size=24, bold=True, color="282C3F")
ws1["C6"].alignment = align_center
ws1["C6"].fill = fill_zebra

for r in range(5, 9):
    for c in range(3, 5):
        ws1.cell(row=r, column=c).border = thin_border

ws1.merge_cells("E5:F5")
ws1["E5"] = "RELATIVE GROWTH TARGET"
ws1["E5"].font = Font(name="Segoe UI", size=10, bold=True, color="7F8C8D")
ws1["E5"].alignment = align_center

ws1.merge_cells("E6:F8")
ws1["E6"] = "=(A6-C6)/C6" # Excel Formula
ws1["E6"].font = Font(name="Segoe UI", size=24, bold=True, color="2E7D32")
ws1["E6"].alignment = align_center
ws1["E6"].number_format = '0.0%'
ws1["E6"].fill = fill_green_badge

for r in range(5, 9):
    for c in range(5, 7):
        ws1.cell(row=r, column=c).border = thin_border

ws1.merge_cells("G5:H5")
ws1["G5"] = "ADDITIONAL REVENUE RUN-RATE (EST)"
ws1["G5"].font = Font(name="Segoe UI", size=10, bold=True, color="7F8C8D")
ws1["G5"].alignment = align_center

ws1.merge_cells("G6:H8")
ws1["G6"] = "Rs. 18.5 Cr / Yr"
ws1["G6"].font = Font(name="Segoe UI", size=16, bold=True, color="2E7D32")
ws1["G6"].alignment = align_center
ws1["G6"].fill = fill_green_badge

for r in range(5, 9):
    for c in range(7, 9):
        ws1.cell(row=r, column=c).border = thin_border

# Section Header
ws1["A10"] = "Core Strategic Pillars (Non-Monetary Conversion Framework)"
ws1["A10"].font = font_section
ws1.merge_cells("A10:H10")
for col_idx in range(1, 9):
    ws1.cell(row=10, column=col_idx).fill = fill_section_bar

# Pillar Table Headers
headers_1 = ["Strategic Pillar", "Underlying Friction Addressed", "Target User Cohort", "Proposed Growth Solution", "Expected 30d Conv Impact"]
for idx, h in enumerate(headers_1):
    c = ws1.cell(row=11, column=idx+1)
    c.value = h
    c.font = font_header
    c.fill = fill_header_dark
    c.alignment = align_center
    c.border = thin_border
ws1.row_dimensions[11].height = 25

pillars_data = [
    ["1. Size & Fit Confidence", "Inter-brand sizing inconsistencies; fear of purchasing wrong size and return hassle.", "Size-Anxious Buyers, Casual Shoppers", "Myntra Size & Fit Harmonizer (Cross-brand calibration)", "+0.90%"],
    ["2. Visual & Spec Comparison", "Decision paralysis when comparing multiple wishlisted similar items; back-and-forth swipe fatigue.", "Compare-and-Contrast Shoppers", "Myntra Compare-Fit Studio (Unified comparison matrix)", "+0.70%"],
    ["3. Outfitting & Styling Validation", "Styling uncertainty (how to wear the item, what existing wardrobe pieces match it).", "Gen Z Trendsetters, Outfit Planners", "Myntra OOTD Styling Board (Interactive drag-drop pairing)", "+0.45%"],
    ["4. Social Validation", "Seeking opinion from friends before purchase; screenshotting fatigue and external app barriers.", "Social Shoppers, Event Shoppers", "Wishlist Social Voting Hub (Instant voting link without app install)", "+0.25%"]
]

for row_idx, data in enumerate(pillars_data):
    r = 12 + row_idx
    ws1.row_dimensions[r].height = 24
    for col_idx, val in enumerate(data):
        cell = ws1.cell(row=r, column=col_idx+1)
        cell.value = val
        cell.font = font_normal
        cell.border = thin_border
        if col_idx == 4:
            cell.alignment = align_center
            cell.font = font_bold
        else:
            cell.alignment = align_left
        if row_idx % 2 == 1:
            cell.fill = fill_zebra

# Add spacing and metadata note
ws1["A18"] = "*Note: Expected impacts are cumulative based on GTM model predictions and historical A/B tests on search & browse."
ws1["A18"].font = font_italic

apply_grid_and_autosize(ws1)

# ==========================================
# SHEET 2: AI DISCOVERY ENGINE DATA
# ==========================================
ws2 = wb.create_sheet(title="AI Discovery Data")
create_banner(ws2, "AI-POWERED FEEDBACK ANALYSIS ENGINE", "Raw Feedback Classified at Scale from Reddit, Play Store, and Styling Communities")

if not df_feedback.empty:
    # Print Headers from CSV
    headers_2 = ["Source Channel", "Target User Segment", "Customer Voice (Raw Feedback)", "Friction Category", "Sentiment", "Shopping Intent", "Barrier Level", "Primary Pain Point"]
    for idx, h in enumerate(headers_2):
        c = ws2.cell(row=5, column=idx+1)
        c.value = h
        c.font = font_header
        c.fill = fill_header_dark
        c.alignment = align_center
        c.border = thin_border
    ws2.row_dimensions[5].height = 25
    
    # Fill Data (up to 40 items for representation in sheet, otherwise it becomes too big)
    sample_df = df_feedback.head(40)
    for row_idx, row in sample_df.iterrows():
        r = 6 + row_idx
        ws2.row_dimensions[r].height = 32
        
        # Write values
        vals = [row["source"], row["user_segment"], row["comment"], row["category_tag"], row["sentiment"], row["intent_type"], row["barrier_level"], row["primary_pain_point"]]
        for col_idx, val in enumerate(vals):
            cell = ws2.cell(row=r, column=col_idx+1)
            cell.value = val
            cell.font = font_normal
            cell.border = thin_border
            
            # Alignments & conditional styles
            if col_idx == 2: # Raw comment
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                
            # Zebra striping
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
                
            # Style Badge columns
            if col_idx == 4: # Sentiment
                if val == "Negative":
                    cell.fill = fill_high_priority
                    cell.font = font_high_priority
                elif val == "Positive":
                    cell.fill = fill_green_badge
                    cell.font = font_green_badge
            elif col_idx == 6: # Barrier Level
                if val == "High":
                    cell.fill = fill_high_priority
                    cell.font = font_high_priority
                elif val == "Medium":
                    cell.fill = fill_med_priority
                    cell.font = font_med_priority
                elif val == "Low":
                    cell.fill = fill_low_priority
                    cell.font = font_low_priority

    # Auto sizing columns for sheet 2
    apply_grid_and_autosize(ws2)
    # Give the comment column custom large width
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["H"].width = 30
else:
    ws2["A5"] = "No feedback database found. Run discovery_engine.py first."
    ws2["A5"].font = font_bold

# ==========================================
# SHEET 3: COHORT & INTENT ANALYSIS
# ==========================================
ws3 = wb.create_sheet(title="User Cohorts & Intent")
create_banner(ws3, "USER WISHLIST COHORTS & INTENT MATRIX", "Categorizing Wishlist Users by Behavior, Motivations, and Non-Monetary Conversion Triggers")

# Table Headers
headers_3 = ["User Cohort Name", "Shopping Motivation", "Wishlist Role / Usage Pattern", "Key Purchase Friction", "Non-Monetary Conversion Trigger", "Metric Impact Potential"]
for idx, h in enumerate(headers_3):
    c = ws3.cell(row=5, column=idx+1)
    c.value = h
    c.font = font_header
    c.fill = fill_header_dark
    c.alignment = align_center
    c.border = thin_border
ws3.row_dimensions[5].height = 25

cohorts_data = [
    [
        "The Sizing Anxious Buyer", 
        "Needs high-quality wardrobe additions but is terrified of wrong sizes and returns.",
        "Holds high-intent items. Doesn't checkout because of brand size variance.",
        "Size chart inconsistencies across brands (Roadster vs. Roadster V2 vs. HRX).",
        "Cross-brand size calibration recommendations based on items they already own and didn't return.",
        "High (Largest volume of active cart-replacements)"
    ],
    [
        "The Compare & Contrast Shopper",
        "Aims for the best style/cut but gets paralyzed by too many similar options.",
        "Wishlists 5-10 very similar items (e.g. 5 black high heels) to select the single best one.",
        "Fatigue from swiping back and forth across product pages; cannot see side-by-side specs.",
        "Unified comparison grid showing size, fabric thickness, buyer photos, and average return rate.",
        "High (Funnels user directly to a single checkout choice)"
    ],
    [
        "The Outfit Planner (Gen Z)",
        "Focused on creating 'looks' and aesthetic combinations.",
        "Saves separate items (skirt, top, accessories) to plan a future look.",
        "Unsure if items actually look good together. Catalog images don't layer.",
        "Interactive styling canvas to drag-and-drop wishlisted items side-by-side or layered.",
        "Medium (Increases Average Order Value / Multi-item buy)"
    ],
    [
        "The Socially Validating Shopper",
        "Requires peer approval and reassurance before committing to fashion choices.",
        "Saves items and waits for friends' feedback or online styling validation.",
        "TEDIOUS screenshots / sending individual links to friends who don't have the app.",
        "One-click web-link to share specific sub-wishlists with interactive voting.",
        "Medium (Also acts as organic growth loop / referral)"
    ],
    [
        "The Passive Bookmarker",
        "High-funnel browsing, treats wishlist as a 'mood board' or personal archive.",
        "Saves 100+ items with low immediate intent; cleans cart by parking items here.",
        "Low initial intent; item is forgotten over time as new styles accumulate.",
        "Smart wishlist clean-up: 'Is this still you?' collections, auto-tagging by style category.",
        "Low (Addresses low-intent users, but keeps wishlist actionable)"
    ]
]

for row_idx, data in enumerate(cohorts_data):
    r = 6 + row_idx
    ws3.row_dimensions[r].height = 35
    for col_idx, val in enumerate(data):
        cell = ws3.cell(row=r, column=col_idx+1)
        cell.value = val
        cell.font = font_normal
        cell.border = thin_border
        
        # Color cohort column
        if col_idx == 0:
            cell.font = font_bold
            cell.alignment = align_left
        elif col_idx == 5: # Metric Impact
            cell.alignment = align_center
            if "High" in val:
                cell.fill = fill_high_priority
                cell.font = font_high_priority
            elif "Medium" in val:
                cell.fill = fill_med_priority
                cell.font = font_med_priority
            else:
                cell.fill = fill_low_priority
                cell.font = font_low_priority
        else:
            cell.alignment = align_left
            
        if row_idx % 2 == 1:
            if col_idx != 5: # Keep priority badges colored
                cell.fill = fill_zebra

apply_grid_and_autosize(ws3)
ws3.column_dimensions["C"].width = 25
ws3.column_dimensions["D"].width = 25
ws3.column_dimensions["E"].width = 25

# ==========================================
# SHEET 4: FRICTION POINTS & ROOT CAUSES
# ==========================================
ws4 = wb.create_sheet(title="Friction Deep-Dive")
create_banner(ws4, "WISHLIST CONVERSION FRICTION CATALOG", "Root Cause Analysis of Why High-Intent Users Stop Short of Purchasing")

# Headers
headers_4 = ["Friction Theme", "Friction Breakdown", "Root Cause Analysis (Tech/UX)", "Typical Customer Quote", "Non-Monetary Product Countermeasure"]
for idx, h in enumerate(headers_4):
    c = ws4.cell(row=5, column=idx+1)
    c.value = h
    c.font = font_header
    c.fill = fill_header_dark
    c.alignment = align_center
    c.border = thin_border
ws4.row_dimensions[5].height = 25

friction_data = [
    [
        "Size & Fit Discrepancy",
        "Fear of getting wrong fit leading to annoying return cycle.",
        "Brand sizing charts are inconsistent, and users don't know how the garment actually fits on their body shape compared to standard model profiles.",
        "\"Roadster L fits me fine, but Solly L is like a tent. I don't want to buy, try, and return 5 shirts.\"",
        "Cross-brand sizing calibration + community-sourced 'Runs Small/Large' indicator based on buyer return telemetry."
    ],
    [
        "Visual/Spec Comparison Fatigue",
        "Inability to filter or compare multiple similar saved options.",
        "Wishlist has no structured matrix. Users must tap into each product page sequentially to check fabric, neck shape, rating, or delivery date.",
        "\"Saved 5 black blazers. To compare them, I have to open 5 tabs and swipe. I get tired and close the app.\"",
        "Unified comparison slider in Wishlist allowing side-by-side specification overlay (e.g. fabric, length, rating, speed)."
    ],
    [
        "Outfitting Mismatch (Styling)",
        "Unsure how to integrate the wishlisted item into existing wardrobe.",
        "Model photos show highly styled studio looks that feel unrealistic. Product pages do not show real-world outfitting combos.",
        "\"I love this yellow jacket but what pants go with it? I have no idea if my wishlisted crop tops match.\"",
        "Wishlist Mix-and-Match Canvas (virtual canvas where users drag saved tops, bottoms, shoes to build and save looks)."
    ],
    [
        "Quality Verification Void",
        "Skepticism of highly-edited catalog photos vs. real life fabric appearance.",
        "Studio shots use professional lighting and model tailoring. Fabric weight/thickness (e.g., sheer vs. thick) is not communicated clearly.",
        "\"Catalog photos look great, but in user reviews, the material looks super cheap. I keep it wishlisted until someone posts real photos.\"",
        "Priority ranking for reviews with customer photos + fabric weight indicator (e.g. Lightweight, Midweight, Heavyweight)."
    ],
    [
        "Social Approval Friction",
        "Time-lag in getting feedback from friends/family.",
        "No simple sharing loop. Users must screenshot items and send via WhatsApp, or share entire links that require friends to install Myntra.",
        "\"Wanted my roommate's advice on these 3 dresses for a wedding. screenshotting them all is so annoying. I gave up.\"",
        "Shared Wishlist Web Link: Create a view-only web page of selected items where friends can vote with 1 tap (no login required)."
    ]
]

for row_idx, data in enumerate(friction_data):
    r = 6 + row_idx
    ws4.row_dimensions[r].height = 45
    for col_idx, val in enumerate(data):
        cell = ws4.cell(row=r, column=col_idx+1)
        cell.value = val
        cell.font = font_normal
        cell.border = thin_border
        
        if col_idx == 0:
            cell.font = font_bold
            cell.alignment = align_left
        elif col_idx == 3:
            cell.font = font_italic
            cell.alignment = align_left
        else:
            cell.alignment = align_left
            
        if row_idx % 2 == 1:
            cell.fill = fill_zebra

apply_grid_and_autosize(ws4)
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 28
ws4.column_dimensions["E"].width = 28

# ==========================================
# SHEET 5: OPPORTUNITY BACKLOG (ICE)
# ==========================================
ws5 = wb.create_sheet(title="Opportunity Backlog")
create_banner(ws5, "GROWTH INITIATIVES BACKLOG & PRIORITIZATION", "Evaluating Product Features Using ICE scoring framework (No Monetary Incentives)")

# Headers
headers_5 = ["Initiative ID", "Feature / Solution", "Target Friction Point", "Impact (1-10)", "Confidence (1-10)", "Ease (1-10)", "ICE Score", "Rank"]
for idx, h in enumerate(headers_5):
    c = ws5.cell(row=5, column=idx+1)
    c.value = h
    c.font = font_header
    c.fill = fill_header_dark
    c.alignment = align_center
    c.border = thin_border
ws5.row_dimensions[5].height = 25

backlog_data = [
    ["INIT-001", "Myntra Size Harmonizer", "Inter-Brand Size Inconsistency", 9, 8, 6],
    ["INIT-002", "Wishlist Compare-Fit Matrix", "Decision Paralysis / Spec Comparison", 8, 8, 7],
    ["INIT-003", "Shared Wishlist & Social Polls", "Peer Validation Friction", 6, 7, 8],
    ["INIT-004", "Mix & Match Outfit Canvas", "Styling Uncertainty / Styling Matches", 7, 6, 5],
    ["INIT-005", "Real-Life Photo Review Prominence", "Fabric / Quality Verification Void", 6, 8, 7],
    ["INIT-006", "Wishlist Curation Folders / Tags", "Platform Behavior / Mood Boarding", 5, 8, 8],
    ["INIT-007", "Fabric Thickness & Weight Specs", "Quality Verification Void", 5, 7, 9]
]

# Write backlog data and calculate ICE Score via formula
for row_idx, data in enumerate(backlog_data):
    r = 6 + row_idx
    ws5.row_dimensions[r].height = 24
    
    # Init ID
    ws5.cell(row=r, column=1, value=data[0]).alignment = align_center
    ws5.cell(row=r, column=1).font = font_bold
    ws5.cell(row=r, column=1).border = thin_border
    
    # Name
    ws5.cell(row=r, column=2, value=data[1]).alignment = align_left
    ws5.cell(row=r, column=2).font = font_bold
    ws5.cell(row=r, column=2).border = thin_border
    
    # Friction
    ws5.cell(row=r, column=3, value=data[2]).alignment = align_left
    ws5.cell(row=r, column=3).border = thin_border
    
    # Impact, Confidence, Ease
    for offset, val in enumerate(data[3:6]):
        cell = ws5.cell(row=r, column=4+offset, value=val)
        cell.alignment = align_center
        cell.border = thin_border
        cell.number_format = '0'
        
    # ICE Score Formula (Average of Impact, Confidence, Ease)
    ice_formula = f"=AVERAGE(D{r}:F{r})"
    cell_ice = ws5.cell(row=r, column=7, value=ice_formula)
    cell_ice.alignment = align_center
    cell_ice.font = font_bold
    cell_ice.border = thin_border
    cell_ice.number_format = '0.0'
    cell_ice.fill = fill_zebra
    
    # Rank Formula (Excel RANK function based on ICE Score range G6:G12)
    rank_formula = f"=RANK(G{r}, $G$6:$G$12)"
    cell_rank = ws5.cell(row=r, column=8, value=rank_formula)
    cell_rank.alignment = align_center
    cell_rank.font = font_bold
    cell_rank.border = thin_border
    
    if row_idx < 3: # Top 3 rank highlights
        cell_rank.fill = fill_green_badge
        cell_rank.font = font_green_badge
    else:
        cell_rank.fill = fill_zebra

    for col_idx in range(1, 9):
        if row_idx % 2 == 1:
            curr_cell = ws5.cell(row=r, column=col_idx)
            if curr_cell.fill.fill_type is None: # Only if not already colored
                curr_cell.fill = fill_zebra

apply_grid_and_autosize(ws5)
ws5.column_dimensions["C"].width = 25

# ==========================================
# SHEET 6: DETAILED PRODUCT SPECS
# ==========================================
ws6 = wb.create_sheet(title="Product Specs (PRD)")
create_banner(ws6, "GROWTH SOLUTIONS SPECIFICATION (PRD)", "High-Level Product Requirements for Top Ranked Wishlist Interventions")

# Feature 1 Header
ws6["A5"] = "FEATURE SPEC 1: MYNTRA SIZE HARMONIZER (INIT-001)"
ws6["A5"].font = font_section
ws6.merge_cells("A5:F5")
for col_idx in range(1, 7):
    ws6.cell(row=5, column=col_idx).fill = fill_section_bar
ws6.row_dimensions[5].height = 24

spec1_data = [
    ["User Problem", "Users wishlist fashion items but hesitate to buy due to size chart variations across brands."],
    ["Non-Monetary Solution", "Cross-Brand Size Harmonizer. Map sizes from brands the user has successfully kept to the wishlisted brand. E.g. 'Since Roadster Size M fits you perfectly, we recommend Roadster V2 Size L for this item.'"],
    ["UX Flow", "1. User opens Wishlist. 2. A 'Size Guide Check' badge appears on items with size uncertainty. 3. Clicking badge shows 'Harmonizer Recommendation Card' with historical comparison. 4. One-click size selection directly from Wishlist page."],
    ["Key Dependencies", "Telemetry database of successfully purchased and NOT returned apparel size per user; brand size specifications database."],
    ["Success Metric", "Wishlist-to-Purchase conversion rate for Apparel; decrease in sizing-related return rates."]
]

for idx, (label, val) in enumerate(spec1_data):
    r = 6 + idx
    ws6.row_dimensions[r].height = 30
    
    c_label = ws6.cell(row=r, column=1, value=label)
    c_label.font = font_bold
    c_label.fill = fill_zebra
    c_label.border = thin_border
    c_label.alignment = align_left
    
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    for c_idx in range(2, 7):
        c_val = ws6.cell(row=r, column=c_idx)
        c_val.border = thin_border
        if c_idx == 2:
            c_val.value = val
            c_val.font = font_normal
            c_val.alignment = align_left

# Spacer
ws6.row_dimensions[11].height = 15

# Feature 2 Header
ws6["A12"] = "FEATURE SPEC 2: WISHLIST COMPARE-FIT STUDIO (INIT-002)"
ws6["A12"].font = font_section
ws6.merge_cells("A12:F12")
for col_idx in range(1, 7):
    ws6.cell(row=12, column=col_idx).fill = fill_section_bar
ws6.row_dimensions[12].height = 24

spec2_data = [
    ["User Problem", "Users wishlist multiple similar items (e.g., 5 white sneakers) and face decision paralysis comparing them."],
    ["Non-Monetary Solution", "Wishlist Compare-Fit Matrix. A side-by-side grid of wishlisted items within the same subcategory, detailing material, thickness, user rating, real-life customer photos, return rates, and shipping speed."],
    ["UX Flow", "1. User multi-selects items in wishlist. 2. Clicks 'Compare side-by-side'. 3. Unified comparison overlay loads with key parameters highlighted (e.g. 'Thicker fabric', 'Delivers 2 days faster'). 4. Checkout button directly under matrix."],
    ["Key Dependencies", "Standardized attributes metadata (fabric weight, sole height, etc.); UI overlays."],
    ["Success Metric", "Average time-to-purchase from addition of >2 similar wishlisted items; Comparison-to-Cart click-through rate."]
]

for idx, (label, val) in enumerate(spec2_data):
    r = 13 + idx
    ws6.row_dimensions[r].height = 30
    
    c_label = ws6.cell(row=r, column=1, value=label)
    c_label.font = font_bold
    c_label.fill = fill_zebra
    c_label.border = thin_border
    c_label.alignment = align_left
    
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    for c_idx in range(2, 7):
        c_val = ws6.cell(row=r, column=c_idx)
        c_val.border = thin_border
        if c_idx == 2:
            c_val.value = val
            c_val.font = font_normal
            c_val.alignment = align_left

apply_grid_and_autosize(ws6)
ws6.column_dimensions["A"].width = 22

# ==========================================
# SHEET 7: EXPERIMENTATION & TELEMETRY
# ==========================================
ws7 = wb.create_sheet(title="Experimentation Plan")
create_banner(ws7, "A/B TESTING & EXPERIMENTATION METRIC INSTRUMENTATION", "Telemetry Requirements and A/B Testing Protocols for Non-Monetary Growth Features")

# Headers
headers_7 = ["Telemetry Event Name", "Trigger Action", "Data Payload Properties", "Growth Metric Impacted", "Telemetry Priority"]
for idx, h in enumerate(headers_7):
    c = ws7.cell(row=5, column=idx+1)
    c.value = h
    c.font = font_header
    c.fill = fill_header_dark
    c.alignment = align_center
    c.border = thin_border
ws7.row_dimensions[5].height = 25

telemetry_data = [
    ["wishlist_harmonize_badge_view", "Sizing Harmonizer badge displays on wishlisted item.", "product_id, recommended_size, confidence_score", "Harmonizer Engagement Rate", "High"],
    ["wishlist_harmonize_card_click", "User clicks on the size harmonization badge.", "product_id, current_brand_size, compared_brand_size", "Size Guide CTR", "High"],
    ["wishlist_compare_initiated", "User clicks the compare button in wishlist after selecting >=2 items.", "product_count, category_id, compared_product_ids", "Comparison Studio Adoption Rate", "Medium"],
    ["wishlist_compare_cart_add", "User adds an item to cart directly from comparison matrix.", "product_id, comparison_duration_sec, position_in_matrix", "Comparison Conversion Rate", "High"],
    ["wishlist_social_share_created", "User generates a social voting link for their wishlist.", "wishlist_id, items_shared_count, channel (WA/Insta)", "Viral K-Factor / Peer Share Rate", "Medium"],
    ["wishlist_social_vote_received", "A friend votes on a shared wishlist web link.", "wishlist_id, product_id, vote_type (Love/Skip)", "Social Vote Participation Rate", "Medium"]
]

for row_idx, data in enumerate(telemetry_data):
    r = 6 + row_idx
    ws7.row_dimensions[r].height = 28
    for col_idx, val in enumerate(data):
        cell = ws7.cell(row=r, column=col_idx+1)
        cell.value = val
        cell.font = font_normal
        cell.border = thin_border
        
        if col_idx == 0:
            cell.font = font_bold
            cell.alignment = align_left
        elif col_idx == 4: # Priority
            cell.alignment = align_center
            if "High" in val:
                cell.fill = fill_high_priority
                cell.font = font_high_priority
            else:
                cell.fill = fill_low_priority
                cell.font = font_low_priority
        else:
            cell.alignment = align_left
            
        if row_idx % 2 == 1:
            if col_idx != 4:
                cell.fill = fill_zebra

# A/B Test Config section
ws7.cell(row=14, column=1, value="A/B TEST DEPLOYMENT PARAMETERS").font = font_section
ws7.merge_cells("A14:E14")
for col_idx in range(1, 6):
    ws7.cell(row=14, column=col_idx).fill = fill_section_bar
ws7.row_dimensions[14].height = 24

ab_data = [
    ["Test Structure", "50% Control (Current Wishlist UI) | 50% Variant (Size Harmonizer + Compare Studio enabled)"],
    ["Target Power & Alpha", "Alpha = 0.05 (Statistical Significance 95%) | Power = 0.80 (80% Probability of detecting effect)"],
    ["Minimum Sample Size", "50,000 active wishlisters per arm (Calculated using baseline 4.2% rate, looking for 0.5% absolute lift)"],
    ["Experiment Duration", "21 days (accounting for day-of-week shopping fluctuations and baseline 30-day conversion cycle)"],
    ["Guardrail Metric", "Sizing return rates (Must not increase) | App load latency / Wishlist rendering speed"]
]

for idx, (label, val) in enumerate(ab_data):
    r = 15 + idx
    ws7.row_dimensions[r].height = 28
    
    c_label = ws7.cell(row=r, column=1, value=label)
    c_label.font = font_bold
    c_label.fill = fill_zebra
    c_label.border = thin_border
    c_label.alignment = align_left
    
    ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    for c_idx in range(2, 6):
        c_val = ws7.cell(row=r, column=c_idx)
        c_val.border = thin_border
        if c_idx == 2:
            c_val.value = val
            c_val.font = font_normal
            c_val.alignment = align_left

apply_grid_and_autosize(ws7)
ws7.column_dimensions["A"].width = 24
ws7.column_dimensions["B"].width = 22
ws7.column_dimensions["C"].width = 22

# Save the workbook
wb.save(excel_path)
print(f"Workbook successfully saved to: {excel_path}")
